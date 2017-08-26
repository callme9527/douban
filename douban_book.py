# -*- coding:utf-8 -*-
__author__ = '9527'
import time
import urllib2
import urllib
import sys
import os
import openpyxl
from getopt import getopt, GetoptError
from bs4 import BeautifulSoup
from random import choice
from os.path import abspath, isfile
from config import template_url, headers, error_urls_file
from openpyxl import Workbook
reload(sys)
sys.setdefaultencoding('utf-8')

tag = 'book'


def get_books(category, error_f):
    error_f.write(category+'\n')
    category = urllib.quote(category)
    error_f.write(category+'\n')
    fail500_times = 0
    fail400_times = 0
    start = 0
    books = []
    while True:
        try:
            url = template_url.format(category, tag) + "?start="
            url = url+str(start)
            print url
            req = urllib2.Request(url, headers=choice(headers))
            res = urllib2.urlopen(req, timeout=2)
            con = res.read()

            soup = BeautifulSoup(con, 'html.parser', from_encoding='utf-8')
            book_list_soup = soup.find('div', {'class': 'mod book-list'}).find_all('dl')
            if not book_list_soup:
                print u'抓取完毕,URL为:', url
                print 'size:', len(books)
                return books
            for book_soup in book_list_soup:
                try:
                    book = {}

                    book_dt_a = book_soup.find('dt').find('a')
                    book['img'] = book_dt_a.find('img').get('src', '')
                    book['url'] = book_dt_a.get('href', '')

                    book_dd = book_soup.find('dd')
                    book['name'] = book_dd.find('a', {'class': 'title'}).string.strip()
                    book_desc = book_dd.find('div', {'class': 'desc'}).string.strip().split('/')
                    book['price'] = book_desc[-1]
                    book['time'] = book_desc[-2]
                    book['pub'] = book_desc[-3]
                    book['country'] = u'中国'
                    author = book_desc[0]
                    if author.startswith('['):
                        index = author.find(']')
                        book['country'] = author[:(index+1)]
                        author = author[(index+1):]
                    book['author'] = ' '.join([author]+book_desc[1:-3]).strip()
                    try:
                        book['rating'] = book_dd.find('span', {'class': 'rating_nums'}).string.strip()
                    except:
                        book['rating'] = 0.0
                    print book
                    books.append(book)
                except Exception, e:
                    error_f.write('[-]ParserError:'+str(e)+' '+url+' book is: '+book['name']+' '+str(book_desc)+'\n')
            start += 15
        except urllib2.URLError, e:
            if hasattr(e, 'code'):
                error_f.write('[-]UrlError: '+url+' '+str(e.reason)+' '+str(e.code)+'\r\n')
                if e.code/100 == 5:
                    fail500_times += 1
                    if fail500_times <= 2:
                        print u'[-] 服务器炸了,休息片刻...'
                        time.sleep(0.5)
                        continue
                    else:
                        print u'[-] 该页面%s一直500，访问下一页面' % url
                if e.code/100 == 4:
                    print u'[-] 抓取URL %s 出错,抓取下一页面...' % url
                    fail400_times += 1
                    if fail400_times >= 5:
                        print u'输入的分类很可能不正确，页面一直无法访问'
                        break
            print e.reason
            start += 15
            error_f.write('[-]UrlError: '+url+' '+str(e.reason)+'\r\n')
            continue
        except Exception, e:
            print str(e)
            start += 15
            error_f.write('[-]Error: '+url+' '+str(e)+'\r\n')
            continue
        except KeyboardInterrupt:
            error_f.close()


def save_as_excel(categorys, xf, ef):
    wb = Workbook()
    if isfile(abspath(xf)):
        wb = openpyxl.load_workbook(xf)
        os.system('del '+xf)
    for category in categorys.split(','):
        ws = wb.create_sheet(title=category)
        books = get_books(category.encode('utf-8'), ef)
        books = sorted(books, key=lambda x: x['rating'], reverse=True)
        ws.append(['书名', '价格', '国家', '作|译者', '出版社', '时间', '总分', '图片', 'URL'])
        for book in books:
            ws.append([book['name'], book['price'], book['country'], book['author'],
                       book['pub'], book['time'], book['rating'], book['img'], book['url']])
    wb.save(xf)
    wb.close()


def usage():
    print u'''          -h[--help]: for help.
          -c[--category] for category[s]. e.g:'小说'，'漫画'.
          -f[--efile] for saving error url.'''
    sys.exit()

if __name__ == '__main__':
    # get_books('小说', f)
    efile = 'error_urls.txt'
    try:
        options, args = getopt(sys.argv[1:], 'c:e:h', ['category=', 'efile=', 'help'])
        for name, value in options:
            if name in ('-h', '--help'):
                usage()
            elif name in ('-c', '--category'):
                categorys = value.decode('gbk')
            elif name in ('-e', '--efile'):
                efile = value.decode('gbk').encode('utf-8')
                if not isfile(abs(efile)):
                    usage()
            else:
                usage()
    except GetoptError, e:
        print str(e)
        usage()

    ef = open(efile, 'w')
    ef.write(categorys.encode('utf-8')+'\n')
    save_as_excel(categorys, 'books.xlsx', ef)
    ef.close()











